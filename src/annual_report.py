#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Relatório Anual de actividade do Conversor Excel → PDF.

Agrega dados do histórico de conversões (tabela ``history``) para um
determinado ano civil e gera relatórios em PDF ou Excel.
"""

import os
from collections import defaultdict
from datetime import datetime

from src.database import _get_connection

_MESES_PT = [
    '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
]


# ============================================
# AGREGAÇÃO DE DADOS
# ============================================

def get_annual_data(year: int) -> dict:
    """Agrega dados do histórico para o ano indicado.

    Args:
        year: Ano civil (ex: 2026).

    Returns:
        Dict com:
        - ``year``            — ano solicitado
        - ``total``           — total de conversões
        - ``success``         — conversões com sucesso
        - ``errors``          — conversões com erro
        - ``clients_total``   — total de clientes processados
        - ``by_month``        — lista de 12 dicts (jan=0 … dez=11) com
                                ``month``, ``label``, ``conversions``,
                                ``clients``, ``success``, ``errors``
        - ``by_mode``         — dict {mode: count}
        - ``top_files``       — lista de top-10 ficheiros (``file``, ``count``)
    """
    conn = _get_connection()
    try:
        rows = conn.execute(
            """SELECT timestamp, source_file, mode, clients_count, success
               FROM history
               WHERE timestamp >= ? AND timestamp < ?
               ORDER BY timestamp""",
            (f"{year}-01-01", f"{year + 1}-01-01"),
        ).fetchall()
    finally:
        conn.close()

    # Estruturas de agregação
    by_month = [
        {'month': m, 'label': _MESES_PT[m],
         'conversions': 0, 'clients': 0, 'success': 0, 'errors': 0}
        for m in range(1, 13)
    ]
    by_mode: dict[str, int] = defaultdict(int)
    file_counts: dict[str, int] = defaultdict(int)
    total = success_count = error_count = clients_total = 0

    for row in rows:
        try:
            ts = datetime.fromisoformat(row['timestamp'])
            m = ts.month
        except (ValueError, TypeError):
            continue

        total += 1
        ok = bool(row['success'])
        clients = row['clients_count'] or 0

        by_month[m - 1]['conversions'] += 1
        by_month[m - 1]['clients'] += clients
        if ok:
            by_month[m - 1]['success'] += 1
            success_count += 1
        else:
            by_month[m - 1]['errors'] += 1
            error_count += 1

        clients_total += clients
        by_mode[row['mode'] or 'desconhecido'] += 1
        file_counts[row['source_file'] or ''] += 1

    top_files = sorted(
        [{'file': f, 'count': c} for f, c in file_counts.items() if f],
        key=lambda x: -x['count'],
    )[:10]

    return {
        'year': year,
        'total': total,
        'success': success_count,
        'errors': error_count,
        'clients_total': clients_total,
        'by_month': by_month,
        'by_mode': dict(by_mode),
        'top_files': top_files,
    }


def get_available_years() -> list[int]:
    """Retorna os anos com registos no histórico, mais recente primeiro."""
    conn = _get_connection()
    try:
        rows = conn.execute(
            "SELECT DISTINCT substr(timestamp, 1, 4) AS yr FROM history ORDER BY yr DESC"
        ).fetchall()
        return [int(r['yr']) for r in rows if r['yr'] and r['yr'].isdigit()]
    finally:
        conn.close()


# ============================================
# EXPORTAÇÃO PDF
# ============================================

def generate_annual_report_pdf(year: int, output_path: str, config: dict) -> str:
    """Gera o relatório anual em PDF.

    Args:
        year:        Ano civil.
        output_path: Caminho do ficheiro PDF a criar.
        config:      Configuração da aplicação.

    Returns:
        Caminho do ficheiro PDF gerado.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    data = get_annual_data(year)
    colors_cfg = config.get('colors', {})
    header_cfg = config.get('header', {})

    empresa = header_cfg.get('company_name', 'Empresa')
    header_bg = colors.HexColor(colors_cfg.get('header_bg', '#2d3748'))
    header_text_color = colors.HexColor(colors_cfg.get('header_text', '#ffffff'))
    row_alt = colors.HexColor(colors_cfg.get('row_alt', '#f7fafc'))
    total_bg = colors.HexColor(colors_cfg.get('total_bg', '#edf2f7'))
    total_text = colors.HexColor(colors_cfg.get('total_text', '#1a365d'))
    border_color = colors.HexColor(colors_cfg.get('border', '#e2e8f0'))

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=20*mm, leftMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
        title=f"Relatório Anual {year}",
        author=empresa,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Title2', parent=styles['Heading1'],
                              fontSize=18, alignment=TA_CENTER, spaceAfter=4))
    styles.add(ParagraphStyle('Sub', parent=styles['Normal'],
                              fontSize=10, alignment=TA_CENTER,
                              textColor=colors.HexColor('#4a5568'), spaceAfter=16))
    styles.add(ParagraphStyle('SH', parent=styles['Heading2'],
                              fontSize=12, spaceBefore=12, spaceAfter=6))
    styles.add(ParagraphStyle('Foot', parent=styles['Normal'],
                              fontSize=8, textColor=colors.gray, alignment=TA_RIGHT))

    def _table(data_rows, col_widths, total_row=False):
        t = Table(data_rows, colWidths=col_widths)
        cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text_color),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1 if not total_row else -2),
             [colors.white, row_alt]),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
        ]
        if total_row and len(data_rows) > 1:
            tr = len(data_rows) - 1
            cmds += [
                ('FONTNAME', (0, tr), (-1, tr), 'Helvetica-Bold'),
                ('BACKGROUND', (0, tr), (-1, tr), total_bg),
                ('TEXTCOLOR', (0, tr), (-1, tr), total_text),
            ]
        t.setStyle(TableStyle(cmds))
        return t

    elements = []

    # --- Título ---
    elements.append(Paragraph(empresa, styles['Title2']))
    elements.append(Paragraph(f"Relatório Anual de Actividade — {year}", styles['Sub']))

    # --- Resumo global ---
    elements.append(Paragraph("Resumo Global", styles['SH']))
    taxa_sucesso = f"{data['success'] / data['total'] * 100:.1f}%" if data['total'] else "—"
    summary_data = [
        ['Indicador', 'Valor'],
        ['Total de Conversões', str(data['total'])],
        ['Conversões com Sucesso', str(data['success'])],
        ['Conversões com Erro', str(data['errors'])],
        ['Taxa de Sucesso', taxa_sucesso],
        ['Total de Clientes Processados', str(data['clients_total'])],
    ]
    elements.append(_table(summary_data, [110*mm, 60*mm]))
    elements.append(Spacer(1, 6*mm))

    # --- Por mês ---
    elements.append(Paragraph("Actividade por Mês", styles['SH']))
    month_rows = [['Mês', 'Conversões', 'Clientes', 'Sucesso', 'Erros']]
    total_conv = total_clients = total_suc = total_err = 0
    for m in data['by_month']:
        month_rows.append([
            m['label'],
            str(m['conversions']) if m['conversions'] else '—',
            str(m['clients']) if m['clients'] else '—',
            str(m['success']) if m['success'] else '—',
            str(m['errors']) if m['errors'] else '—',
        ])
        total_conv += m['conversions']
        total_clients += m['clients']
        total_suc += m['success']
        total_err += m['errors']
    month_rows.append(['TOTAL', str(total_conv), str(total_clients),
                       str(total_suc), str(total_err)])
    elements.append(_table(month_rows, [50*mm, 35*mm, 35*mm, 35*mm, 15*mm], total_row=True))
    elements.append(Spacer(1, 6*mm))

    # --- Por modo ---
    if data['by_mode']:
        elements.append(Paragraph("Conversões por Modo", styles['SH']))
        mode_rows = [['Modo', 'Nº de Conversões']]
        for mode, count in sorted(data['by_mode'].items(), key=lambda x: -x[1]):
            mode_rows.append([mode, str(count)])
        elements.append(_table(mode_rows, [110*mm, 60*mm]))
        elements.append(Spacer(1, 6*mm))

    # --- Top ficheiros ---
    if data['top_files']:
        elements.append(Paragraph("Ficheiros Mais Processados", styles['SH']))
        file_rows = [['Ficheiro', 'Conversões']]
        for tf in data['top_files']:
            file_rows.append([tf['file'], str(tf['count'])])
        elements.append(_table(file_rows, [130*mm, 40*mm]))
        elements.append(Spacer(1, 6*mm))

    # --- Rodapé ---
    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph(
        f"Gerado a {datetime.now().strftime('%d/%m/%Y às %H:%M')} · Conversor Excel PDF",
        styles['Foot'],
    ))

    doc.build(elements)
    return output_path


# ============================================
# EXPORTAÇÃO EXCEL
# ============================================

def generate_annual_report_excel(year: int, output_path: str) -> str:
    """Gera o relatório anual em Excel (.xlsx).

    Args:
        year:        Ano civil.
        output_path: Caminho do ficheiro .xlsx a criar.

    Returns:
        Caminho do ficheiro gerado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_annual_data(year)

    wb = Workbook()

    # Estilos reutilizáveis
    h_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    h_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    alt_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    total_fill = PatternFill(start_color='EDF2F7', end_color='EDF2F7', fill_type='solid')
    total_font = Font(name='Calibri', size=10, bold=True, color='1A365D')
    normal_font = Font(name='Calibri', size=9)
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center')
    right = Alignment(horizontal='right')

    def _header_row(ws, row, values):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = h_font
            c.fill = h_fill
            c.alignment = center
            c.border = border

    def _data_row(ws, row, values, is_total=False, is_alt=False):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = total_font if is_total else normal_font
            if is_total:
                c.fill = total_fill
            elif is_alt:
                c.fill = alt_fill
            c.border = border
            if col > 1:
                c.alignment = right

    # --- Folha 1: Resumo ---
    ws1 = wb.active
    ws1.title = f"Resumo {year}"
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 20

    ws1.cell(row=1, column=1, value=f"Relatório Anual {year}").font = Font(
        name='Calibri', size=14, bold=True)
    ws1.merge_cells('A1:B1')

    _header_row(ws1, 3, ['Indicador', 'Valor'])
    taxa = f"{data['success'] / data['total'] * 100:.1f}%" if data['total'] else '—'
    summary = [
        ('Total de Conversões', data['total']),
        ('Conversões com Sucesso', data['success']),
        ('Conversões com Erro', data['errors']),
        ('Taxa de Sucesso', taxa),
        ('Total de Clientes Processados', data['clients_total']),
    ]
    for i, (label, val) in enumerate(summary, 4):
        _data_row(ws1, i, [label, val], is_alt=(i % 2 == 1))

    # --- Folha 2: Por Mês ---
    ws2 = wb.create_sheet(title="Por Mês")
    for col, w in zip('ABCDE', [20, 15, 12, 12, 10]):
        ws2.column_dimensions[col].width = w
    _header_row(ws2, 1, ['Mês', 'Conversões', 'Clientes', 'Sucesso', 'Erros'])
    t_conv = t_cli = t_suc = t_err = 0
    for i, m in enumerate(data['by_month'], 2):
        _data_row(ws2, i,
                  [m['label'], m['conversions'], m['clients'], m['success'], m['errors']],
                  is_alt=(i % 2 == 0))
        t_conv += m['conversions']
        t_cli += m['clients']
        t_suc += m['success']
        t_err += m['errors']
    _data_row(ws2, 14, ['TOTAL', t_conv, t_cli, t_suc, t_err], is_total=True)

    # --- Folha 3: Por Modo ---
    if data['by_mode']:
        ws3 = wb.create_sheet(title="Por Modo")
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 18
        _header_row(ws3, 1, ['Modo', 'Conversões'])
        for i, (mode, count) in enumerate(
            sorted(data['by_mode'].items(), key=lambda x: -x[1]), 2
        ):
            _data_row(ws3, i, [mode, count], is_alt=(i % 2 == 0))

    # --- Folha 4: Top Ficheiros ---
    if data['top_files']:
        ws4 = wb.create_sheet(title="Top Ficheiros")
        ws4.column_dimensions['A'].width = 45
        ws4.column_dimensions['B'].width = 15
        _header_row(ws4, 1, ['Ficheiro', 'Conversões'])
        for i, tf in enumerate(data['top_files'], 2):
            _data_row(ws4, i, [tf['file'], tf['count']], is_alt=(i % 2 == 0))

    wb.save(output_path)
    return output_path
