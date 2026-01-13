#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo conversor Excel → PDF.
Classe principal para conversão de ficheiros Excel para PDF formatado.
"""

import os
from datetime import datetime

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, LETTER, A3
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER

from src.config import DEFAULT_CONFIG
class ExcelToPDFConverter:
    """Classe para converter dados de Excel para PDF formatado."""
    
    PAGE_SIZES = {
        'A4': A4,
        'A3': A3,
        'Letter': LETTER,
    }
    
    def __init__(self, excel_path: str, output_pdf_path: str = None, config: dict = None):
        self.excel_path = excel_path
        self.config = config or DEFAULT_CONFIG
        
        # Determinar output_path baseado na configuração, por default é o mesmo do excel_path
        if output_pdf_path:
            self.output_pdf_path = output_pdf_path
        else:
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            output_folder = self.config['output'].get('output_folder', '')
            if not output_folder:
                output_folder = os.path.dirname(excel_path)
            
            if self.config['output'].get('add_timestamp', False):
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                self.output_pdf_path = os.path.join(output_folder, f"{base_name}_{timestamp}.pdf")
            else:
                self.output_pdf_path = os.path.join(output_folder, f"{base_name}.pdf")
        
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configura estilos personalizados."""
        colors_cfg = self.config['colors']
        
        self.styles.add(ParagraphStyle(
            name='CompanyTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor(colors_cfg['title']),
            alignment=TA_CENTER,
            spaceAfter=6
        ))
        
        self.styles.add(ParagraphStyle(
            name='SubTitle',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#4a5568'),
            alignment=TA_CENTER,
            spaceAfter=12
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2d3748'),
            spaceBefore=12,
            spaceAfter=6,
        ))
        
        self.styles.add(ParagraphStyle(
            name='NormalText',
            parent=self.styles['Normal'],
            fontSize=self.config['table']['font_size'],
            leading=12
        ))
        
        self.styles.add(ParagraphStyle(
            name='Footer',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#718096'),
            alignment=TA_CENTER
        ))

    def read_excel_data(self) -> dict:
        """Lê os dados do ficheiro Excel."""
        # Tentar carregar com valores calculados primeiro, depois com fórmulas como fallback
        try:
            wb = load_workbook(self.excel_path, data_only=True)
        except Exception:
            wb = load_workbook(self.excel_path)
        
        data = {
            # Configurável
            'empresa': {},
            'cliente': {},
            'documento': {},
            'itens': [],
            'observacoes': '',
            'mes_referencia': '',
            'tipo_relatorio': 'MAPA DE CONTABILIDADE'
        }
        
        # Ler folha de configuração
        if 'Configuracao' in wb.sheetnames:
            ws_config = wb['Configuracao']
            for row in ws_config.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    campo = str(row[0]).strip().lower()
                    valor = str(row[1]).strip() if row[1] else ''
                    
                    if campo in ['nome_empresa', 'morada_empresa', 'telefone_empresa', 'email_empresa', 'nif_empresa']:
                        data['empresa'][campo.replace('_empresa', '')] = valor
                    elif campo in ['nome_cliente', 'morada_cliente', 'telefone_cliente', 'nif_cliente']:
                        data['cliente'][campo.replace('_cliente', '')] = valor
                    elif campo in ['numero_documento', 'data_documento', 'tipo_documento']:
                        data['documento'][campo.replace('_documento', '')] = valor
                    elif campo == 'observacoes':
                        data['observacoes'] = valor
        
        # Usar dados do header config se não existirem no Excel
        header_cfg = self.config['header']
        if not data['empresa'].get('nome'):
            data['empresa']['nome'] = header_cfg.get('company_name', '')
        if not data['empresa'].get('morada'):
            data['empresa']['morada'] = header_cfg.get('company_address', '')
        if not data['empresa'].get('telefone'):
            data['empresa']['telefone'] = header_cfg.get('company_phone', '')
        if not data['empresa'].get('email'):
            data['empresa']['email'] = header_cfg.get('company_email', '')
        if not data['empresa'].get('nif'):
            data['empresa']['nif'] = header_cfg.get('company_nif', '')
        
        # Ler folha de itens (primeira folha activa ou específica)
        ws_itens = wb.active
        for sheet_name in ['Folha1', 'Sheet1', 'Itens', 'Pecas', 'Dados', 'Contas']:
            if sheet_name in wb.sheetnames:
                ws_itens = wb[sheet_name]
                break
        
        # Definir colunas de interesse para formato de contabilidade
        # Mapa de colunas: índice -> nome normalizado
        colunas_contabilidade = {
            'nr.': 'Nr.',
            'nr': 'Nr.',
            'cliente': 'Cliente',
            'contab': 'CONTAB',
            'iva': 'Iva',
            'subtotal': 'Subtotal',
            'extras': 'Extras',
            'duodécimos': 'Duodécimos',
            'duodecimos': 'Duodécimos',
            's.social ger': 'S.Social GER',
            's.social': 'S.Social GER',
            's.soc emp': 'S.Soc Emp',
            'ret. irs': 'Ret. IRS',
            'ret.irs': 'Ret. IRS',
            'ret. irs ext': 'Ret. IRS EXT',
            'sbtx/fcomp': 'SbTx/Fcomp',
            'sbtx': 'SbTx/Fcomp',
            'outro': 'Outro',
            'total': 'TOTAL',
            'nif': 'NIF',
            'sigla': 'SIGLA',
            'mês': 'Mês',
            'mes': 'Mês',
            'data': 'Data',
        }
        
        # Colunas a incluir no PDF (ordem desejada)
        colunas_pdf = ['Nr.', 'SIGLA', 'Cliente', 'CONTAB', 'Iva', 'Subtotal', 
                       'Extras', 'Duodécimos', 'S.Social GER', 'S.Soc Emp', 
                       'Ret. IRS', 'Ret. IRS EXT', 'SbTx/Fcomp', 'Outro', 'TOTAL']
        
        # Encontrar cabeçalhos - procurar linha com palavras-chave de contabilidade
        headers = []
        header_indices = {}  # mapeia nome normalizado -> índice da coluna
        header_row = 1
        contab_found_idx = -1  # Para identificar a área correta de colunas
        
        for row_num, row in enumerate(ws_itens.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any(cell for cell in row):
                row_text = ' '.join(str(c).lower() for c in row if c)
                # Detectar linha de cabeçalhos de contabilidade
                if any(kw in row_text for kw in ['nr.', 'cliente', 'contab', 'total', 'iva', 'subtotal', 'sigla']):
                    # Primeiro passo: encontrar a posição de CONTAB para saber a área correta
                    for i, cell in enumerate(row):
                        if cell and str(cell).strip().lower() == 'contab':
                            contab_found_idx = i
                            break
                    
                    # Segundo passo: mapear todas as colunas
                    for i, cell in enumerate(row):
                        if cell:
                            cell_text = str(cell).strip()
                            cell_lower = cell_text.lower()
                            
                            # Match exato primeiro para colunas com nomes semelhantes
                            if cell_lower == 'ret. irs ext':
                                if 'Ret. IRS EXT' in colunas_pdf:
                                    header_indices['Ret. IRS EXT'] = i
                                continue
                            elif cell_lower == 'ret. irs':
                                if 'Ret. IRS' in colunas_pdf:
                                    header_indices['Ret. IRS'] = i
                                continue
                            elif cell_lower == 'total':
                                # Preferir TOTAL depois de CONTAB (área de contabilidade mensal)
                                if contab_found_idx > 0 and i > contab_found_idx:
                                    if 'TOTAL' in colunas_pdf:
                                        header_indices['TOTAL'] = i
                                elif 'TOTAL' not in header_indices:
                                    if 'TOTAL' in colunas_pdf:
                                        header_indices['TOTAL'] = i
                                continue
                            
                            # Mapear usando dicionário para outras colunas
                            for key, normalized in colunas_contabilidade.items():
                                if key == cell_lower or cell_lower.startswith(key):
                                    # Só adicionar se estiver na lista de colunas PDF
                                    if normalized in colunas_pdf or normalized in ['Mês', 'Data', 'NIF']:
                                        # Evitar sobrescrever se já mapeado
                                        if normalized not in header_indices:
                                            header_indices[normalized] = i
                                    break
                    header_row = row_num
                    break
        
        # Se não encontrou cabeçalhos de contabilidade, tentar formato genérico
        if not header_indices:
            for row_num, row in enumerate(ws_itens.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                if row and any(cell for cell in row):
                    row_text = ' '.join(str(c).lower() for c in row if c)
                    if any(kw in row_text for kw in ['codigo', 'código', 'designacao', 'designação', 'quantidade', 'qtd', 'peça', 'peca', 'ref', 'descri']):
                        headers = [str(c).strip() if c else f'Col{i}' for i, c in enumerate(row)]
                        header_row = row_num
                        break
            
            if not headers:
                headers = ['Código', 'Designação', 'Quantidade', 'Preço Unit.', 'Total']
                header_row = 1
        
        # Capturar mês de referência da primeira linha de dados
        mes_referencia = None
        
        # Ler dados
        for row in ws_itens.iter_rows(min_row=header_row + 1, values_only=True):
            if row and any(cell for cell in row):
                # Verificar se é uma linha vazia (apenas None ou strings vazias)
                values_in_row = [cell for cell in row if cell is not None and str(cell).strip() != '']
                if not values_in_row:
                    continue
                
                item = {}
                
                if header_indices:
                    # Formato contabilidade
                    for col_name, col_idx in header_indices.items():
                        if col_idx < len(row):
                            value = row[col_idx]
                            # Converter valores
                            if value is not None:
                                if isinstance(value, (int, float)):
                                    item[col_name] = value
                                elif isinstance(value, str) and value.startswith('='):
                                    # Fórmula não calculada - tentar 0
                                    item[col_name] = 0
                                else:
                                    item[col_name] = value
                            else:
                                item[col_name] = ''
                    
                    # Capturar mês de referência
                    if not mes_referencia and 'Mês' in item and item['Mês']:
                        mes_referencia = str(item['Mês']).strip()
                        data['mes_referencia'] = mes_referencia
                    
                    # Verificar se tem dados relevantes (Nr. ou Cliente)
                    if item.get('Nr.') or item.get('Cliente'):
                        data['itens'].append(item)
                else:
                    # Formato genérico
                    for i, header in enumerate(headers):
                        if i < len(row):
                            item[header] = row[i] if row[i] is not None else ''
                    if any(v for v in item.values()):
                        data['itens'].append(item)
        
        wb.close()
        return data

    def create_header(self, data: dict) -> list:
        """Cria o cabeçalho do documento."""
        elements = []
        
        if not self.config['header'].get('show_header', True):
            return elements
        
        empresa = data.get('empresa', {})
        
        # Logo (se existir)
        logo_path = self.config['header'].get('logo_path', '')
        if logo_path and os.path.exists(logo_path):
            try:
                img = Image(logo_path, width=50*mm, height=20*mm)
                elements.append(img)
                elements.append(Spacer(1, 5*mm))
            except Exception:
                pass
        
        # Título da empresa
        nome = empresa.get('nome', 'EMPRESA')
        elements.append(Paragraph(nome, self.styles['CompanyTitle']))
        
        # Informações de contacto
        morada = empresa.get('morada', '')
        telefone = empresa.get('telefone', '')
        email = empresa.get('email', '')
        nif = empresa.get('nif', '')
        
        info_parts = []
        if morada:
            info_parts.append(morada)
        contact_line = []
        if telefone:
            contact_line.append(f"Tel: {telefone}")
        if email:
            contact_line.append(f"Email: {email}")
        if nif:
            contact_line.append(f"NIF: {nif}")
        
        if info_parts or contact_line:
            info_text = "<br/>".join(info_parts + [" | ".join(contact_line)])
            elements.append(Paragraph(info_text, self.styles['SubTitle']))
        
        elements.append(Spacer(1, 8*mm))
        
        return elements

    def create_document_info(self, data: dict) -> list:
        """Cria a secção de informações do documento."""
        elements = []
        
        doc_info = data.get('documento', {})
        cliente = data.get('cliente', {})
        
        tipo_doc = doc_info.get('tipo', 'GUIA DE REMESSA')
        num_doc = doc_info.get('numero', datetime.now().strftime('GR%Y%m%d-001'))
        data_doc = doc_info.get('data', datetime.now().strftime('%d/%m/%Y'))
        
        doc_data = [
            [Paragraph(f"<b>{tipo_doc}</b>", self.styles['SectionHeader']), 
             '', 
             Paragraph("<b>CLIENTE</b>", self.styles['SectionHeader'])],
            [f"Nº: {num_doc}", '', f"Nome: {cliente.get('nome', '-')}"],
            [f"Data: {data_doc}", '', f"Morada: {cliente.get('morada', '-')}"],
            ['', '', f"NIF: {cliente.get('nif', '-')}"],
        ]
        
        doc_table = Table(doc_data, colWidths=[70*mm, 20*mm, 90*mm])
        doc_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        elements.append(doc_table)
        elements.append(Spacer(1, 8*mm))
        
        return elements

    def create_items_table(self, data: dict) -> list:
        """Cria a tabela de itens."""
        elements = []
        
        itens = data.get('itens', [])
        
        if not itens:
            elements.append(Paragraph("Sem itens para apresentar.", self.styles['NormalText']))
            return elements
        
        # Verificar se é formato de contabilidade
        primeiro_item = itens[0] if itens else {}
        is_contabilidade = any(key in primeiro_item for key in ['Nr.', 'Cliente', 'CONTAB', 'TOTAL', 'SIGLA'])
        
        if is_contabilidade:
            # Obter colunas da configuração
            contab_cfg = self.config.get('contabilidade', {})
            colunas_str = contab_cfg.get('colunas', 'Nr., SIGLA, Cliente, CONTAB, Iva, Subtotal, Extras, Duodécimos, S.Social GER, S.Soc Emp, Ret. IRS, Ret. IRS EXT, SbTx/Fcomp, Outro, TOTAL')
            colunas_ordem = [c.strip() for c in colunas_str.split(',')]
            
            # Filtrar apenas colunas que existem nos dados
            headers = [col for col in colunas_ordem if any(col in item for item in itens)]
            
            # Nomes abreviados para cabeçalhos (caber melhor na tabela)
            header_display = {
                'Nr.': 'Nr.',
                'SIGLA': 'Sigla',
                'Cliente': 'Cliente',
                'CONTAB': 'Contab.',
                'Iva': 'IVA',
                'Subtotal': 'Subtotal',
                'Extras': 'Extras',
                'Duodécimos': 'Duod.',
                'S.Social GER': 'SS Ger.',
                'S.Soc Emp': 'SS Emp.',
                'Ret. IRS': 'Ret.IRS',
                'Ret. IRS EXT': 'IRS Ext.',
                'SbTx/Fcomp': 'SbTx',
                'Outro': 'Outro',
                'TOTAL': 'TOTAL',
            }
            
            # Formatar valores monetários
            def format_value(val, col_name):
                if val is None or val == '':
                    return ''
                if isinstance(val, (int, float)):
                    # Colunas numéricas - formatar como número com 2 casas
                    if col_name in ['Nr.']:
                        return str(int(val)) if val else ''
                    elif col_name in ['CONTAB', 'Iva', 'Subtotal', 'Extras', 'Duodécimos', 
                                     'S.Social GER', 'S.Soc Emp', 'Ret. IRS', 'Ret. IRS EXT',
                                     'SbTx/Fcomp', 'Outro', 'TOTAL']:
                        if val == 0:
                            return ''
                        return f"{val:.2f}€"
                    return str(val)
                return str(val)
            
            # Criar dados da tabela com nomes abreviados
            display_headers = [header_display.get(h, h) for h in headers]
            table_data = [display_headers]
            for item in itens:
                row = [format_value(item.get(h, ''), h) for h in headers]
                table_data.append(row)
            
            # Calcular larguras específicas para contabilidade
            # Em landscape A4: ~277mm de largura útil
            available_width = 267 * mm
            col_widths = []
            
            for h in headers:
                if h == 'Nr.':
                    col_widths.append(9 * mm)
                elif h == 'SIGLA':
                    col_widths.append(14 * mm)
                elif h == 'Cliente':
                    col_widths.append(50 * mm)  # Coluna mais larga
                elif h == 'TOTAL':
                    col_widths.append(16 * mm)
                else:
                    # Colunas numéricas
                    col_widths.append(13 * mm)
            
            # Ajustar para caber na largura disponível
            total = sum(col_widths)
            if total > available_width:
                col_widths = [w * (available_width / total) for w in col_widths]
            
            # Título da tabela
            mes_ref = data.get('mes_referencia', '')
            titulo_tabela = f"<b>MAPA DE CONTABILIDADE - {mes_ref}</b>" if mes_ref else "<b>MAPA DE CONTABILIDADE</b>"
        
        else:
            # Formato genérico (original)
            all_keys = set()
            for item in itens:
                all_keys.update(item.keys())
            
            priority = ['Código', 'Ref', 'Referência', 'Designação', 'Descrição', 'Quantidade', 'Qtd', 'Preço Unit.', 'Preço', 'Total']
            headers = sorted(list(all_keys), key=lambda x: priority.index(x) if x in priority else 999)
            
            # Criar dados da tabela
            table_data = [headers]
            for item in itens:
                row = [str(item.get(h, '')) for h in headers]
                table_data.append(row)
            
            # Calcular larguras
            num_cols = len(headers)
            available_width = 180 * mm
            col_widths = [available_width / num_cols] * num_cols
            
            for i, h in enumerate(headers):
                h_lower = h.lower()
                if 'designa' in h_lower or 'descri' in h_lower:
                    col_widths[i] = available_width * 0.35
                elif 'codigo' in h_lower or 'ref' in h_lower:
                    col_widths[i] = available_width * 0.12
            
            total = sum(col_widths)
            col_widths = [w * (available_width / total) for w in col_widths]
            
            titulo_tabela = "<b>LISTA DE PEÇAS / ITENS</b>"
        
        # Estilos da tabela
        colors_cfg = self.config['colors']
        table_cfg = self.config['table']
        contab_cfg = self.config.get('contabilidade', {})
        
        # Usar font size menor para contabilidade (muitas colunas)
        font_size = 7 if is_contabilidade else table_cfg['font_size']
        header_font_size = 8 if is_contabilidade else table_cfg['header_font_size']
        row_padding = 4 if is_contabilidade else table_cfg['row_padding']
        
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(colors_cfg['header_bg'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(colors_cfg['header_text'])),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), row_padding + 3),
            ('TOPPADDING', (0, 0), (-1, 0), row_padding + 3),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('BOTTOMPADDING', (0, 1), (-1, -1), row_padding),
            ('TOPPADDING', (0, 1), (-1, -1), row_padding),
        ]
        
        if is_contabilidade:
            # Alinhar colunas numéricas à direita, texto à esquerda
            for i, h in enumerate(headers):
                if h in ['Nr.', 'CONTAB', 'Iva', 'Subtotal', 'Extras', 'Duodécimos', 
                        'S.Social GER', 'S.Soc Emp', 'Ret. IRS', 'Ret. IRS EXT',
                        'SbTx/Fcomp', 'Outro', 'TOTAL']:
                    style_cmds.append(('ALIGN', (i, 1), (i, -1), 'RIGHT'))
                else:
                    style_cmds.append(('ALIGN', (i, 1), (i, -1), 'LEFT'))
            
            # Destacar coluna TOTAL com cor de fundo
            total_idx = headers.index('TOTAL') if 'TOTAL' in headers else -1
            if total_idx >= 0 and contab_cfg.get('destacar_total', True):
                style_cmds.append(('FONTNAME', (total_idx, 0), (total_idx, -1), 'Helvetica-Bold'))
                style_cmds.append(('BACKGROUND', (total_idx, 1), (total_idx, -1), 
                                  colors.HexColor(colors_cfg.get('total_bg', '#edf2f7'))))
                style_cmds.append(('TEXTCOLOR', (total_idx, 1), (total_idx, -1), 
                                  colors.HexColor(colors_cfg.get('total_text', '#1a365d'))))
            
            # Agrupar colunas por cor de fundo para facilitar leitura
            # Grupo 1: Nr., SIGLA, Cliente (sem cor)
            # Grupo 2: CONTAB, Iva, Subtotal (cor suave)
            # Grupo 3: Extras, Duodécimos, SS, etc. (sem cor)
            # Grupo 4: TOTAL (destacado)
            grupo2_cols = ['CONTAB', 'Iva', 'Subtotal']
            for i, h in enumerate(headers):
                if h in grupo2_cols:
                    style_cmds.append(('BACKGROUND', (i, 0), (i, 0), colors.HexColor('#3182ce')))
            
        else:
            style_cmds.append(('ALIGN', (0, 1), (-1, -1), 'LEFT'))
        
        if table_cfg.get('alternate_rows', True):
            style_cmds.append(('ROWBACKGROUNDS', (0, 1), (-1, -1), 
                             [colors.white, colors.HexColor(colors_cfg['row_alt'])]))
        
        if table_cfg.get('show_grid', True):
            style_cmds.append(('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(colors_cfg['border'])))
            style_cmds.append(('BOX', (0, 0), (-1, -1), 1, colors.HexColor(colors_cfg['border'])))
        
        items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        items_table.setStyle(TableStyle(style_cmds))
        
        elements.append(Paragraph(titulo_tabela, self.styles['SectionHeader']))
        elements.append(Spacer(1, 3*mm))
        elements.append(items_table)
        elements.append(Spacer(1, 8*mm))
        
        return elements

    def create_footer(self, data: dict) -> list:
        """Cria o rodapé."""
        elements = []
        footer_cfg = self.config['footer']
        
        # Observações
        if footer_cfg.get('show_observations', True):
            obs = data.get('observacoes', '')
            if obs:
                elements.append(Paragraph("<b>OBSERVAÇÕES:</b>", self.styles['SectionHeader']))
                elements.append(Paragraph(obs, self.styles['NormalText']))
                elements.append(Spacer(1, 8*mm))
        
        # Assinaturas
        if footer_cfg.get('show_signatures', True):
            sig_data = [
                ['', ''],
                ['_' * 30, '_' * 30],
                ['Emitido por', 'Recebido por'],
                ['Data: ___/___/_____', 'Data: ___/___/_____'],
            ]
            
            sig_table = Table(sig_data, colWidths=[85*mm, 85*mm])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            elements.append(Spacer(1, 10*mm))
            elements.append(sig_table)
        
        # Rodapé customizado
        custom_footer = footer_cfg.get('custom_footer', '')
        if custom_footer:
            elements.append(Spacer(1, 5*mm))
            elements.append(Paragraph(custom_footer, self.styles['Footer']))
        
        # Data de geração
        if footer_cfg.get('show_date', True):
            elements.append(Spacer(1, 5*mm))
            footer_text = f"Documento gerado a {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
            elements.append(Paragraph(footer_text, self.styles['Footer']))
        
        return elements

    def generate_pdf(self) -> str:
        """Gera o PDF."""
        data = self.read_excel_data()
        
        # Verificar se é formato de contabilidade
        primeiro_item = data.get('itens', [{}])[0] if data.get('itens') else {}
        is_contabilidade = any(key in primeiro_item for key in ['Nr.', 'Cliente', 'CONTAB', 'TOTAL', 'SIGLA'])
        
        # Configurar página
        pdf_cfg = self.config['pdf']
        page_size = self.PAGE_SIZES.get(pdf_cfg['page_size'], A4)
        
        # Para contabilidade, forçar landscape e margens menores
        if is_contabilidade:
            page_size = (page_size[1], page_size[0])  # Landscape
            margins = {'top': 10, 'bottom': 10, 'left': 10, 'right': 10}
        elif pdf_cfg.get('orientation') == 'landscape':
            page_size = (page_size[1], page_size[0])
            margins = {
                'top': pdf_cfg['margin_top'],
                'bottom': pdf_cfg['margin_bottom'],
                'left': pdf_cfg['margin_left'],
                'right': pdf_cfg['margin_right']
            }
        else:
            margins = {
                'top': pdf_cfg['margin_top'],
                'bottom': pdf_cfg['margin_bottom'],
                'left': pdf_cfg['margin_left'],
                'right': pdf_cfg['margin_right']
            }
        
        doc = SimpleDocTemplate(
            self.output_pdf_path,
            pagesize=page_size,
            rightMargin=margins['right']*mm,
            leftMargin=margins['left']*mm,
            topMargin=margins['top']*mm,
            bottomMargin=margins['bottom']*mm
        )
        
        elements = []
        
        if is_contabilidade:
            # Layout simplificado para contabilidade
            elements.extend(self.create_header(data))
            elements.extend(self.create_items_table(data))
            
            # Rodapé com dados bancários para contabilidade
            elements.append(Spacer(1, 10*mm))
            
            # Dados bancários (substitui "Verificado por" e "Data")
            mes_ref = data.get('mes_referencia', '')
            banking_cfg = self.config.get('banking', {})
            
            if banking_cfg.get('show_banking', True):
                bank_name = banking_cfg.get('bank_name', 'ABANCA')
                iban = banking_cfg.get('iban', 'PT50 0170 3782 0304 0053 5672 9')
                
                banking_text = f"""<b>Nossos Dados Bancários:</b><br/>
                {bank_name}<br/>
                IBAN: {iban}"""
                elements.append(Paragraph(banking_text, self.styles['NormalText']))
                elements.append(Spacer(1, 5*mm))
            
            # Mês de referência
            if mes_ref:
                elements.append(Paragraph(f"Mês de Referência: {mes_ref}", self.styles['NormalText']))
                elements.append(Spacer(1, 8*mm))
            
            # Data de geração
            footer_text = f"Documento gerado a {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
            elements.append(Paragraph(footer_text, self.styles['Footer']))
        else:
            elements.extend(self.create_header(data))
            elements.extend(self.create_document_info(data))
            elements.extend(self.create_items_table(data))
            elements.extend(self.create_footer(data))
        
        doc.build(elements)
        
        return self.output_pdf_path

    def generate_individual_pdfs(self, output_folder: str = None) -> list:
        """Gera um PDF individual para cada cliente/linha do Excel."""
        data = self.read_excel_data()
        itens = data.get('itens', [])
        mes_ref = data.get('mes_referencia', 'SemMes')
        
        if not itens:
            return []
        
        # Criar pasta de destino
        if output_folder is None:
            base_folder = os.path.dirname(self.excel_path)
            output_folder = os.path.join(base_folder, f'PDFs_{mes_ref}')
        
        os.makedirs(output_folder, exist_ok=True)
        
        # Mapeamento de colunas Excel → nomes no PDF
        campo_labels = {
            'CONTAB': 'Serviços de Contabilidade',
            'Iva': 'IVA 23%',
            'Extras': 'Extras',
            'Duodécimos': 'Duodécimos (Despesas Anuais)',
            'S.Social GER': 'Segurança Social Gerentes',
            'S.Soc Emp': 'Segurança Social Empregados',
            'Ret. IRS': 'IRS Retenções Dependentes',
            'Ret. IRS EXT': 'Retenções Indep/Prediais',
            'SbTx/Fcomp': 'Subsídio Férias/Compensação',
            'Outro': 'Outros',
            'TOTAL': 'TOTAL A PAGAR',
        }
        
        # Ordem dos campos
        campos_ordem = ['CONTAB', 'Iva', 'Extras', 'Duodécimos', 'S.Social GER', 
                       'S.Soc Emp', 'Ret. IRS', 'Ret. IRS EXT', 'SbTx/Fcomp', 'Outro', 'TOTAL']
        
        generated_files = []
        
        for item in itens:
            nr = item.get('Nr.', '')
            sigla = item.get('SIGLA', '')
            cliente = item.get('Cliente', '')
            
            if not nr and not cliente:
                continue
            
            # Nome do ficheiro
            filename = f"{nr}_{sigla}.pdf" if sigla else f"{nr}_{cliente[:20]}.pdf"
            filename = filename.replace(' ', '_').replace('/', '-')
            pdf_path = os.path.join(output_folder, filename)
            
            # Gerar PDF individual
            self._create_client_pdf(pdf_path, item, campo_labels, campos_ordem, mes_ref, data)
            generated_files.append(pdf_path)
        
        return generated_files
    
    def _create_client_pdf(self, pdf_path: str, item: dict, campo_labels: dict, 
                           campos_ordem: list, mes_ref: str, data: dict):
        """Cria um PDF individual para um cliente."""
        from reportlab.lib.pagesizes import A4
        
        # Configurar página
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=15*mm,
            bottomMargin=15*mm
        )
        
        elements = []
        colors_cfg = self.config['colors']
        
        # === CABEÇALHO DA EMPRESA ===
        elements.extend(self.create_header(data))
        
        # === DADOS DO CLIENTE ===
        nr = item.get('Nr.', '')
        sigla = item.get('SIGLA', '')
        cliente = item.get('Cliente', '')
        nif = item.get('NIF', '')
        
        elements.append(Spacer(1, 5*mm))
        
        client_info = f"<b>Cliente:</b> {cliente}"
        if sigla:
            client_info += f" ({sigla})"
        elements.append(Paragraph(client_info, self.styles['SectionHeader']))
        
        client_details = f"<b>Nr.:</b> {nr}"
        if nif:
            client_details += f" &nbsp;&nbsp;&nbsp; <b>NIF:</b> {nif}"
        client_details += f" &nbsp;&nbsp;&nbsp; <b>Mês:</b> {mes_ref}"
        elements.append(Paragraph(client_details, self.styles['NormalText']))
        elements.append(Spacer(1, 8*mm))
        
        # === TABELA DE VALORES ===
        elements.append(Paragraph("<b>MAPA DE CONTABILIDADE</b>", self.styles['SectionHeader']))
        elements.append(Spacer(1, 3*mm))
        
        # Construir dados da tabela
        table_data = [['Descrição', 'Valor']]
        
        for campo in campos_ordem:
            if campo in item:
                label = campo_labels.get(campo, campo)
                valor = item.get(campo, 0)
                
                # Formatar valor
                if isinstance(valor, (int, float)):
                    if valor == 0:
                        valor_str = '-'
                    else:
                        valor_str = f"{valor:.2f}€"
                else:
                    valor_str = str(valor) if valor else '-'
                
                table_data.append([label, valor_str])
        
        # Criar tabela
        values_table = Table(table_data, colWidths=[120*mm, 40*mm])
        
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(colors_cfg['header_bg'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(colors_cfg['header_text'])),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(colors_cfg['border'])),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor(colors_cfg['border'])),
        ]
        
        # Destacar linha do TOTAL
        total_row = len(table_data) - 1
        style_cmds.append(('FONTNAME', (0, total_row), (-1, total_row), 'Helvetica-Bold'))
        style_cmds.append(('BACKGROUND', (0, total_row), (-1, total_row), 
                          colors.HexColor(colors_cfg.get('total_bg', '#edf2f7'))))
        style_cmds.append(('TEXTCOLOR', (0, total_row), (-1, total_row), 
                          colors.HexColor(colors_cfg.get('total_text', '#1a365d'))))
        
        # Linhas alternadas
        if self.config['table'].get('alternate_rows', True):
            style_cmds.append(('ROWBACKGROUNDS', (0, 1), (-1, -2), 
                             [colors.white, colors.HexColor(colors_cfg['row_alt'])]))
        
        values_table.setStyle(TableStyle(style_cmds))
        elements.append(values_table)
        
        # === DADOS BANCÁRIOS ===
        elements.append(Spacer(1, 15*mm))
        
        banking_cfg = self.config.get('banking', {})
        
        if banking_cfg.get('show_banking', True):
            bank_name = banking_cfg.get('bank_name', 'ABANCA')
            iban = banking_cfg.get('iban', 'PT50 0170 3782 0304 0053 5672 9')
            
            banking_text = f"""<b>Nossos Dados Bancários:</b><br/>
            {bank_name}<br/>
            IBAN: {iban}"""
            elements.append(Paragraph(banking_text, self.styles['NormalText']))
            elements.append(Spacer(1, 5*mm))
        
        # Mês de referência
        elements.append(Paragraph(f"Mês de Referência: {mes_ref}", self.styles['NormalText']))
        elements.append(Spacer(1, 8*mm))
        
        # Data de geração
        footer_text = f"Documento gerado a {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        elements.append(Paragraph(footer_text, self.styles['Footer']))
        
        doc.build(elements)
