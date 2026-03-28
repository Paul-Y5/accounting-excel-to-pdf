#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo de exportação para Excel formatado.
Gera ficheiros .xlsx prontos a imprimir com formatação profissional.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(data: dict, output_path: str, config: dict) -> str:
    """Exporta dados para Excel formatado.

    Args:
        data: Dados lidos do Excel original (via converter.read_excel_data()).
        output_path: Caminho do ficheiro de saída.
        config: Configurações da aplicação.

    Returns:
        Caminho do ficheiro gerado.
    """
    itens = data.get('itens', [])
    if not itens:
        raise ValueError("Sem dados para exportar.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa Contabilidade"

    # Cores da config
    colors_cfg = config.get('colors', {})
    header_bg_hex = colors_cfg.get('header_bg', '#2d3748').lstrip('#')
    header_text_hex = colors_cfg.get('header_text', '#ffffff').lstrip('#')
    row_alt_hex = colors_cfg.get('row_alt', '#f7fafc').lstrip('#')
    total_bg_hex = colors_cfg.get('total_bg', '#edf2f7').lstrip('#')
    total_text_hex = colors_cfg.get('total_text', '#1a365d').lstrip('#')

    # Estilos
    header_font = Font(name='Calibri', size=10, bold=True, color=header_text_hex)
    header_fill = PatternFill(start_color=header_bg_hex, end_color=header_bg_hex, fill_type='solid')
    alt_fill = PatternFill(start_color=row_alt_hex, end_color=row_alt_hex, fill_type='solid')
    total_fill = PatternFill(start_color=total_bg_hex, end_color=total_bg_hex, fill_type='solid')
    total_font = Font(name='Calibri', size=10, bold=True, color=total_text_hex)
    normal_font = Font(name='Calibri', size=9)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Obter colunas da configuração
    contab_cfg = config.get('contabilidade', {})
    colunas_str = contab_cfg.get('colunas', 'Nr., SIGLA, Cliente, CONTAB, Iva, Subtotal, Extras, Duodécimos, S.Social GER, S.Soc Emp, Ret. IRS, Ret. IRS EXT, SbTx/Fcomp, Outro, TOTAL')
    colunas_ordem = [c.strip() for c in colunas_str.split(',')]

    # Filtrar colunas que existem nos dados
    headers = [col for col in colunas_ordem if any(col in item for item in itens)]

    # Colunas numéricas (para formatação)
    numeric_cols = {'CONTAB', 'Iva', 'Subtotal', 'Extras', 'Duodécimos',
                    'S.Social GER', 'S.Soc Emp', 'Ret. IRS', 'Ret. IRS EXT',
                    'SbTx/Fcomp', 'Outro', 'TOTAL'}

    # === CABEÇALHO DA EMPRESA ===
    empresa = data.get('empresa', {})
    header_cfg = config.get('header', {})
    nome_empresa = empresa.get('nome') or header_cfg.get('company_name', '')
    mes_ref = data.get('mes_referencia', '')

    title_font = Font(name='Calibri', size=14, bold=True)
    subtitle_font = Font(name='Calibri', size=11, color='4A5568')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=nome_empresa).font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    titulo_mapa = f"MAPA DE CONTABILIDADE - {mes_ref}" if mes_ref else "MAPA DE CONTABILIDADE"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=titulo_mapa).font = subtitle_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    # Linha em branco
    start_row = 4

    # === CABEÇALHOS DA TABELA ===
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # === DADOS ===
    for row_idx, item in enumerate(itens, start_row + 1):
        for col_idx, col_name in enumerate(headers, 1):
            value = item.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name in numeric_cols and isinstance(value, (int, float)):
                if value != 0:
                    cell.value = value
                    cell.number_format = '#,##0.00€'
                else:
                    cell.value = None
                cell.alignment = Alignment(horizontal='right')
            elif col_name == 'Nr.' and isinstance(value, (int, float)):
                cell.value = int(value)
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.value = str(value) if value else ''
                cell.alignment = Alignment(horizontal='left')

            cell.font = normal_font
            cell.border = thin_border

            # Cor alternada
            if (row_idx - start_row) % 2 == 0:
                cell.fill = alt_fill

            # Destacar coluna TOTAL
            if col_name == 'TOTAL':
                cell.fill = total_fill
                cell.font = total_font

    # === LARGURAS DAS COLUNAS ===
    for col_idx, col_name in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if col_name == 'Cliente':
            ws.column_dimensions[col_letter].width = 35
        elif col_name in ['Nr.', 'SIGLA']:
            ws.column_dimensions[col_letter].width = 10
        elif col_name == 'TOTAL':
            ws.column_dimensions[col_letter].width = 14
        else:
            ws.column_dimensions[col_letter].width = 12

    # === RODAPÉ: DADOS BANCÁRIOS ===
    last_data_row = start_row + len(itens) + 1
    banking_cfg = config.get('banking', {})
    if banking_cfg.get('show_banking', True):
        bank_name = banking_cfg.get('bank_name', '')
        iban = banking_cfg.get('iban', '')

        ws.cell(row=last_data_row + 1, column=1, value='Dados Bancários:').font = Font(name='Calibri', size=9, bold=True)
        ws.cell(row=last_data_row + 2, column=1, value=f'{bank_name} — IBAN: {iban}').font = Font(name='Calibri', size=9, color='4A5568')

    # === DATA DE GERAÇÃO ===
    footer_row = last_data_row + 4
    ws.cell(row=footer_row, column=1,
            value=f'Gerado a {datetime.now().strftime("%d/%m/%Y às %H:%M")}').font = Font(name='Calibri', size=8, color='718096')

    # Configurar impressão
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f'{start_row}:{start_row}'

    wb.save(output_path)
    return output_path
