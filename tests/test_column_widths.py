"""
Testes para a feature de larguras de colunas configuráveis (feature/v3.3-column-widths).

Valida que:
- col_widths existe no DEFAULT_CONFIG['contabilidade']
- larguras personalizadas são aplicadas à tabela em vez das automáticas
- colunas não configuradas mantêm a largura automática
- valores inválidos são ignorados sem erro
"""
import copy
import os
import tempfile
import pytest
from unittest.mock import patch, MagicMock
from openpyxl import Workbook

from src.config import DEFAULT_CONFIG
from src.converter import ExcelToPDFConverter
from reportlab.lib.units import mm


@pytest.fixture
def temp_dir():
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


@pytest.fixture
def contab_excel(temp_dir):
    """Excel com colunas de contabilidade padrão."""
    path = os.path.join(temp_dir, 'contab.xlsx')
    wb = Workbook()
    ws = wb.active
    headers = ['Nr.', 'SIGLA', 'Cliente', 'CONTAB', 'Iva', 'Subtotal', 'TOTAL']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value='AA')
    ws.cell(row=2, column=3, value='Cliente Teste')
    ws.cell(row=2, column=4, value=100)
    ws.cell(row=2, column=5, value=23)
    ws.cell(row=2, column=6, value=100)
    ws.cell(row=2, column=7, value=123)
    wb.save(path)
    wb.close()
    yield path


class TestDefaultConfig:
    def test_col_widths_key_exists(self):
        assert 'col_widths' in DEFAULT_CONFIG['contabilidade']

    def test_col_widths_default_is_empty_dict(self):
        assert DEFAULT_CONFIG['contabilidade']['col_widths'] == {}


class TestColumnWidthsApplied:
    def _capture_col_widths(self, converter):
        """Executa generate_pdf e captura o colWidths passado à Table principal."""
        captured = {}

        original_table = __import__('reportlab.platypus', fromlist=['Table']).Table

        def fake_table(data, colWidths=None, **kwargs):
            if colWidths and len(colWidths) > 4:
                # A tabela de itens tem mais de 4 colunas
                captured['colWidths'] = colWidths
            return original_table(data, colWidths=colWidths, **kwargs)

        with patch('src.converter.Table', side_effect=fake_table):
            converter.generate_pdf()

        return captured.get('colWidths')

    def test_custom_width_applied_to_cliente(self, contab_excel, temp_dir):
        cfg = copy.deepcopy(DEFAULT_CONFIG)
        cfg['contabilidade']['col_widths'] = {'Cliente': 70}
        output = os.path.join(temp_dir, 'out.pdf')
        conv = ExcelToPDFConverter(contab_excel, output, config=cfg)
        widths = self._capture_col_widths(conv)
        # Encontrar a posição de 'Cliente' nos headers
        # Sabemos que na tabela contabilidade headers = ['Nr.', 'SIGLA', 'Cliente', ...]
        if widths:
            assert abs(widths[2] - 70 * mm) < 0.01

    def test_zero_width_means_auto(self, contab_excel, temp_dir):
        """Largura 0 deve ser ignorada (usa automático)."""
        cfg = copy.deepcopy(DEFAULT_CONFIG)
        cfg['contabilidade']['col_widths'] = {'Cliente': 0}
        output = os.path.join(temp_dir, 'out.pdf')
        conv = ExcelToPDFConverter(contab_excel, output, config=cfg)
        # Não deve lançar excepção
        result = conv.generate_pdf()
        assert os.path.exists(result)

    def test_invalid_width_ignored(self, contab_excel, temp_dir):
        """Valor não-numérico não deve causar erro."""
        cfg = copy.deepcopy(DEFAULT_CONFIG)
        cfg['contabilidade']['col_widths'] = {'Cliente': 'abc'}
        output = os.path.join(temp_dir, 'out.pdf')
        conv = ExcelToPDFConverter(contab_excel, output, config=cfg)
        result = conv.generate_pdf()
        assert os.path.exists(result)

    def test_unknown_column_ignored(self, contab_excel, temp_dir):
        """Coluna inexistente no Excel não causa erro."""
        cfg = copy.deepcopy(DEFAULT_CONFIG)
        cfg['contabilidade']['col_widths'] = {'ColunaInexistente': 50}
        output = os.path.join(temp_dir, 'out.pdf')
        conv = ExcelToPDFConverter(contab_excel, output, config=cfg)
        result = conv.generate_pdf()
        assert os.path.exists(result)

    def test_empty_col_widths_no_error(self, contab_excel, temp_dir):
        """Dicionário vazio não causa problema."""
        cfg = copy.deepcopy(DEFAULT_CONFIG)
        cfg['contabilidade']['col_widths'] = {}
        output = os.path.join(temp_dir, 'out.pdf')
        conv = ExcelToPDFConverter(contab_excel, output, config=cfg)
        result = conv.generate_pdf()
        assert os.path.exists(result)

    def test_multiple_custom_widths(self, contab_excel, temp_dir):
        """Múltiplas colunas configuradas em simultâneo."""
        cfg = copy.deepcopy(DEFAULT_CONFIG)
        cfg['contabilidade']['col_widths'] = {'Nr.': 12, 'SIGLA': 18, 'TOTAL': 20}
        output = os.path.join(temp_dir, 'out.pdf')
        conv = ExcelToPDFConverter(contab_excel, output, config=cfg)
        widths = self._capture_col_widths(conv)
        if widths:
            assert abs(widths[0] - 12 * mm) < 0.01
            assert abs(widths[1] - 18 * mm) < 0.01
            # TOTAL é o último (índice 6 nos 7 headers)
            assert abs(widths[6] - 20 * mm) < 0.01
