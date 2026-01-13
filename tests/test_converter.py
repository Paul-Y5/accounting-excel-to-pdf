"""
Testes unitários para o módulo conversor.
"""

import os
import pytest
import tempfile
from openpyxl import Workbook

from src.config import DEFAULT_CONFIG
from src.converter import ExcelToPDFConverter


@pytest.fixture
def sample_excel_file(temp_dir):
    """Cria um ficheiro Excel de exemplo para testes."""
    excel_path = os.path.join(temp_dir, 'test_contas.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contas'
    
    # Cabeçalhos
    headers = ['Nr.', 'SIGLA', 'Cliente', 'CONTAB', 'Iva', 'TOTAL']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dados de exemplo
    data = [
        [1, 'ABC', 'Cliente ABC', 100.00, 23.00, 123.00],
        [2, 'XYZ', 'Cliente XYZ', 200.00, 46.00, 246.00],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(excel_path)
    wb.close()
    
    yield excel_path
    
    # Limpar
    if os.path.exists(excel_path):
        os.remove(excel_path)


@pytest.fixture
def temp_dir():
    """Cria um diretório temporário."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


class TestExcelToPDFConverter:
    """Testes para a classe ExcelToPDFConverter."""
    
    def test_converter_initialization(self, sample_excel_file):
        """Verifica inicialização do conversor."""
        converter = ExcelToPDFConverter(sample_excel_file)
        
        assert converter.excel_path == sample_excel_file
        assert converter.config is not None
        assert converter.output_pdf_path.endswith('.pdf')
    
    def test_converter_with_custom_config(self, sample_excel_file):
        """Verifica inicialização com configuração customizada."""
        custom_config = DEFAULT_CONFIG.copy()
        custom_config['banking']['bank_name'] = 'CGD'
        
        converter = ExcelToPDFConverter(sample_excel_file, config=custom_config)
        
        assert converter.config['banking']['bank_name'] == 'CGD'
    
    def test_read_excel_data_returns_dict(self, sample_excel_file):
        """Verifica que read_excel_data retorna um dicionário."""
        converter = ExcelToPDFConverter(sample_excel_file)
        data = converter.read_excel_data()
        
        assert isinstance(data, dict)
        assert 'itens' in data
        assert 'empresa' in data
    
    def test_read_excel_data_parses_items(self, sample_excel_file):
        """Verifica que read_excel_data lê os itens corretamente."""
        converter = ExcelToPDFConverter(sample_excel_file)
        data = converter.read_excel_data()
        
        assert len(data['itens']) == 2
        assert data['itens'][0].get('Cliente') == 'Cliente ABC'
    
    def test_generate_pdf_creates_file(self, sample_excel_file, temp_dir):
        """Verifica que generate_pdf cria o ficheiro PDF."""
        output_path = os.path.join(temp_dir, 'output.pdf')
        converter = ExcelToPDFConverter(sample_excel_file, output_path)
        
        result = converter.generate_pdf()
        
        assert os.path.exists(result)
        assert result.endswith('.pdf')
    
    def test_generate_individual_pdfs(self, sample_excel_file, temp_dir):
        """Verifica que generate_individual_pdfs cria múltiplos PDFs."""
        converter = ExcelToPDFConverter(sample_excel_file)
        
        output_folder = os.path.join(temp_dir, 'pdfs_individuais')
        result = converter.generate_individual_pdfs(output_folder)
        
        assert len(result) == 2
        for pdf_path in result:
            assert os.path.exists(pdf_path)


class TestBankingDataInPDF:
    """Testes para verificar dados bancários no PDF."""
    
    def test_banking_config_is_used(self, sample_excel_file):
        """Verifica que a configuração bancária é usada."""
        custom_config = DEFAULT_CONFIG.copy()
        custom_config['banking'] = {
            'show_banking': True,
            'bank_name': 'BANCO TESTE',
            'iban': 'PT50 9999 9999 9999 9999 9999 9',
        }
        
        converter = ExcelToPDFConverter(sample_excel_file, config=custom_config)
        
        assert converter.config['banking']['bank_name'] == 'BANCO TESTE'
