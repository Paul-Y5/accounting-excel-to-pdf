"""
Testes unitários para o módulo de exportação Excel.
"""

import copy
import os
import pytest

from openpyxl import load_workbook

from src.config import DEFAULT_CONFIG
from src.excel_exporter import export_to_excel


@pytest.fixture
def sample_data():
    """Dados de exemplo para exportação."""
    return {
        'empresa': {'nome': 'Empresa Teste, Lda'},
        'mes_referencia': 'Janeiro 2025',
        'itens': [
            {'Nr.': 1, 'SIGLA': 'ET', 'Cliente': 'Empresa Teste', 'CONTAB': 100.0,
             'Iva': 23.0, 'Subtotal': 123.0, 'Extras': 0, 'Duodécimos': 0,
             'S.Social GER': 0, 'S.Soc Emp': 0, 'Ret. IRS': 0, 'Ret. IRS EXT': 0,
             'SbTx/Fcomp': 0, 'Outro': 0, 'TOTAL': 123.0},
            {'Nr.': 2, 'SIGLA': 'AB', 'Cliente': 'ABANCA', 'CONTAB': 200.0,
             'Iva': 46.0, 'Subtotal': 246.0, 'Extras': 10.0, 'Duodécimos': 0,
             'S.Social GER': 0, 'S.Soc Emp': 0, 'Ret. IRS': 0, 'Ret. IRS EXT': 0,
             'SbTx/Fcomp': 0, 'Outro': 0, 'TOTAL': 256.0},
        ],
    }


@pytest.fixture
def sample_config():
    """Configuração de exemplo."""
    return copy.deepcopy(DEFAULT_CONFIG)


class TestExportToExcel:
    """Testes para a função export_to_excel."""

    def test_creates_file(self, tmp_path, sample_data, sample_config):
        """Verifica que o ficheiro Excel é criado."""
        output = str(tmp_path / 'output.xlsx')
        result = export_to_excel(sample_data, output, sample_config)
        assert os.path.exists(result)

    def test_returns_output_path(self, tmp_path, sample_data, sample_config):
        """Verifica que retorna o caminho do ficheiro."""
        output = str(tmp_path / 'output.xlsx')
        result = export_to_excel(sample_data, output, sample_config)
        assert result == output

    def test_raises_on_empty_data(self, tmp_path, sample_config):
        """Verifica que lança erro com dados vazios."""
        output = str(tmp_path / 'output.xlsx')
        with pytest.raises(ValueError, match="Sem dados"):
            export_to_excel({'itens': []}, output, sample_config)

    def test_raises_on_missing_itens(self, tmp_path, sample_config):
        """Verifica que lança erro quando não há itens."""
        output = str(tmp_path / 'output.xlsx')
        with pytest.raises(ValueError, match="Sem dados"):
            export_to_excel({}, output, sample_config)

    def test_sheet_name(self, tmp_path, sample_data, sample_config):
        """Verifica o nome da folha."""
        output = str(tmp_path / 'output.xlsx')
        export_to_excel(sample_data, output, sample_config)
        wb = load_workbook(output)
        assert wb.active.title == "Mapa Contabilidade"
        wb.close()

    def test_company_name_in_header(self, tmp_path, sample_data, sample_config):
        """Verifica que o nome da empresa aparece no cabeçalho."""
        output = str(tmp_path / 'output.xlsx')
        export_to_excel(sample_data, output, sample_config)
        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == 'Empresa Teste, Lda'
        wb.close()

    def test_month_reference_in_title(self, tmp_path, sample_data, sample_config):
        """Verifica que o mês de referência aparece no título."""
        output = str(tmp_path / 'output.xlsx')
        export_to_excel(sample_data, output, sample_config)
        wb = load_workbook(output)
        ws = wb.active
        assert 'Janeiro 2025' in ws.cell(row=2, column=1).value
        wb.close()

    def test_data_rows_present(self, tmp_path, sample_data, sample_config):
        """Verifica que os dados são escritos no ficheiro."""
        output = str(tmp_path / 'output.xlsx')
        export_to_excel(sample_data, output, sample_config)
        wb = load_workbook(output)
        ws = wb.active
        # Verificar que existem valores nas linhas de dados (row 5 e 6, pois 4 é header)
        values = []
        for row in ws.iter_rows(min_row=5, max_row=6, values_only=True):
            values.extend([v for v in row if v is not None])
        assert len(values) > 0
        wb.close()

    def test_banking_data_in_footer(self, tmp_path, sample_data, sample_config):
        """Verifica que os dados bancários aparecem no rodapé."""
        output = str(tmp_path / 'output.xlsx')
        export_to_excel(sample_data, output, sample_config)
        wb = load_workbook(output)
        ws = wb.active
        # Procurar 'Dados Bancários' nas células
        found = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and 'Dados Bancários' in str(cell):
                    found = True
                    break
        assert found, "Dados bancários não encontrados no rodapé"
        wb.close()

    def test_banking_hidden_when_disabled(self, tmp_path, sample_data, sample_config):
        """Verifica que dados bancários não aparecem quando desativados."""
        sample_config['banking']['show_banking'] = False
        output = str(tmp_path / 'output.xlsx')
        export_to_excel(sample_data, output, sample_config)
        wb = load_workbook(output)
        ws = wb.active
        found = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and 'Dados Bancários' in str(cell):
                    found = True
                    break
        assert not found, "Dados bancários não deviam aparecer"
        wb.close()

    def test_without_month_reference(self, tmp_path, sample_config):
        """Verifica exportação sem mês de referência."""
        data = {
            'empresa': {'nome': 'Teste'},
            'mes_referencia': '',
            'itens': [{'Nr.': 1, 'Cliente': 'A', 'TOTAL': 10.0}],
        }
        output = str(tmp_path / 'output.xlsx')
        export_to_excel(data, output, sample_config)
        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "MAPA DE CONTABILIDADE"
        wb.close()

    def test_uses_config_company_name_fallback(self, tmp_path, sample_config):
        """Verifica que usa company_name da config quando empresa não tem nome."""
        data = {
            'empresa': {},
            'mes_referencia': '',
            'itens': [{'Nr.': 1, 'Cliente': 'A', 'TOTAL': 10.0}],
        }
        output = str(tmp_path / 'output.xlsx')
        export_to_excel(data, output, sample_config)
        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == sample_config['header']['company_name']
        wb.close()
