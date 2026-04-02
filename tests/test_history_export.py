"""
Testes unitários para exportação do histórico (CSV e Excel).
"""

import csv
import os
import pytest
from openpyxl import load_workbook

from src import database as db
from src import history


@pytest.fixture(autouse=True)
def isolated_db(tmp_path, monkeypatch):
    """Base de dados isolada com 3 entradas pré-criadas."""
    db_path = str(tmp_path / 'test.db')
    monkeypatch.setattr('src.database._get_db_path', lambda: db_path)
    db.init_db()
    db.add_history_entry('/path/a.xlsx', '/out/a.pdf', 'aggregate', 5, True)
    db.add_history_entry('/path/b.xlsx', '/out/b.pdf', 'individual', 3, True)
    db.add_history_entry('/path/c.xlsx', '', 'aggregate', 0, False, 'Erro de leitura')


# ============================================
# CSV
# ============================================

class TestExportHistoryCsv:
    """Testes para history.export_to_csv."""

    def test_creates_file(self, tmp_path):
        """Verifica que o ficheiro CSV é criado."""
        output = str(tmp_path / 'history.csv')
        result = history.export_to_csv(output)
        assert os.path.exists(result)

    def test_returns_output_path(self, tmp_path):
        """Retorna o caminho do ficheiro gerado."""
        output = str(tmp_path / 'history.csv')
        result = history.export_to_csv(output)
        assert result == output

    def test_csv_has_all_entries(self, tmp_path):
        """CSV contém todas as entradas do histórico."""
        output = str(tmp_path / 'history.csv')
        history.export_to_csv(output)
        with open(output, newline='', encoding='utf-8') as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 3

    def test_csv_has_header_row(self, tmp_path):
        """CSV tem linha de cabeçalho com nomes de colunas."""
        output = str(tmp_path / 'history.csv')
        history.export_to_csv(output)
        with open(output, newline='', encoding='utf-8') as f:
            header = next(csv.reader(f))
        assert 'source_file' in header

    def test_csv_limit_parameter(self, tmp_path):
        """Parâmetro limit restringe o número de linhas exportadas."""
        output = str(tmp_path / 'history.csv')
        history.export_to_csv(output, limit=2)
        with open(output, newline='', encoding='utf-8') as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 2

    def test_csv_contains_source_file(self, tmp_path):
        """CSV contém o nome do ficheiro de origem."""
        output = str(tmp_path / 'history.csv')
        history.export_to_csv(output)
        with open(output, newline='', encoding='utf-8') as f:
            rows = list(csv.DictReader(f))
        source_files = {r['source_file'] for r in rows}
        assert 'a.xlsx' in source_files

    def test_csv_empty_history(self, tmp_path, monkeypatch):
        """Histórico vazio gera ficheiro sem dados (mas sem erro)."""
        db_path = str(tmp_path / 'empty.db')
        monkeypatch.setattr('src.database._get_db_path', lambda: db_path)
        db.init_db()
        output = str(tmp_path / 'empty.csv')
        history.export_to_csv(output)
        assert os.path.exists(output)


# ============================================
# EXCEL
# ============================================

class TestExportHistoryExcel:
    """Testes para history.export_to_excel."""

    def test_creates_file(self, tmp_path):
        """Verifica que o ficheiro Excel é criado."""
        output = str(tmp_path / 'history.xlsx')
        result = history.export_to_excel(output)
        assert os.path.exists(result)

    def test_returns_output_path(self, tmp_path):
        """Retorna o caminho do ficheiro gerado."""
        output = str(tmp_path / 'history.xlsx')
        result = history.export_to_excel(output)
        assert result == output

    def test_sheet_name(self, tmp_path):
        """Folha tem o nome 'Histórico'."""
        output = str(tmp_path / 'history.xlsx')
        history.export_to_excel(output)
        wb = load_workbook(output)
        assert 'Histórico' in wb.sheetnames
        wb.close()

    def test_has_header_row(self, tmp_path):
        """Primeira linha contém cabeçalhos preenchidos."""
        output = str(tmp_path / 'history.xlsx')
        history.export_to_excel(output)
        wb = load_workbook(output)
        ws = wb.active
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert any(h for h in header if h)
        wb.close()

    def test_data_rows_count(self, tmp_path):
        """Número de linhas de dados corresponde ao histórico."""
        output = str(tmp_path / 'history.xlsx')
        history.export_to_excel(output)
        wb = load_workbook(output)
        ws = wb.active
        data_rows = ws.max_row - 1  # excluir cabeçalho
        assert data_rows == 3
        wb.close()

    def test_limit_parameter(self, tmp_path):
        """Parâmetro limit restringe linhas exportadas."""
        output = str(tmp_path / 'history.xlsx')
        history.export_to_excel(output, limit=1)
        wb = load_workbook(output)
        ws = wb.active
        data_rows = ws.max_row - 1
        assert data_rows == 1
        wb.close()

    def test_header_is_bold(self, tmp_path):
        """Cabeçalho está em negrito."""
        output = str(tmp_path / 'history.xlsx')
        history.export_to_excel(output)
        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold is True
        wb.close()

    def test_excel_empty_history(self, tmp_path, monkeypatch):
        """Histórico vazio gera ficheiro Excel válido."""
        db_path = str(tmp_path / 'empty.db')
        monkeypatch.setattr('src.database._get_db_path', lambda: db_path)
        db.init_db()
        output = str(tmp_path / 'empty.xlsx')
        history.export_to_excel(output)
        assert os.path.exists(output)
