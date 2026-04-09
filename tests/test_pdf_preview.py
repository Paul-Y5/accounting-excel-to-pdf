"""
Testes para o módulo de pré-visualização de PDF (feature/v3.3-remaining).

Valida que:
- render_page devolve imagem PIL e contagem de páginas
- página inválida lança IndexError
- ficheiro inexistente lança FileNotFoundError
- get_page_count devolve número correto
"""
import os
import tempfile
import copy
import pytest

from src.config import DEFAULT_CONFIG


@pytest.fixture
def temp_dir():
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


@pytest.fixture
def sample_pdf(temp_dir):
    """Gera um PDF de teste usando o converter."""
    from openpyxl import Workbook
    from src.converter import ExcelToPDFConverter

    excel_path = os.path.join(temp_dir, 'test.xlsx')
    wb = Workbook()
    ws = wb.active
    for c, h in enumerate(['Nr.', 'SIGLA', 'Cliente', 'TOTAL'], 1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value='AA')
    ws.cell(row=2, column=3, value='Teste')
    ws.cell(row=2, column=4, value=100)
    wb.save(excel_path)
    wb.close()

    output = os.path.join(temp_dir, 'test.pdf')
    conv = ExcelToPDFConverter(excel_path, output, config=copy.deepcopy(DEFAULT_CONFIG))
    conv.generate_pdf()
    return output


class TestRenderPage:
    def test_renders_first_page(self, sample_pdf):
        from src.pdf_preview import render_page
        img, total = render_page(sample_pdf, page=0)
        assert img is not None
        assert img.width > 0
        assert img.height > 0
        assert total >= 1

    def test_returns_correct_page_count(self, sample_pdf):
        from src.pdf_preview import render_page
        _, total = render_page(sample_pdf, page=0)
        assert total >= 1

    def test_custom_dpi(self, sample_pdf):
        from src.pdf_preview import render_page
        img_low, _ = render_page(sample_pdf, page=0, dpi=72)
        img_high, _ = render_page(sample_pdf, page=0, dpi=150)
        assert img_high.width > img_low.width

    def test_invalid_page_raises(self, sample_pdf):
        from src.pdf_preview import render_page
        with pytest.raises(IndexError):
            render_page(sample_pdf, page=999)

    def test_negative_page_raises(self, sample_pdf):
        from src.pdf_preview import render_page
        with pytest.raises(IndexError):
            render_page(sample_pdf, page=-1)

    def test_file_not_found(self):
        from src.pdf_preview import render_page
        with pytest.raises(FileNotFoundError):
            render_page('/caminho/inexistente.pdf')


class TestGetPageCount:
    def test_count_matches_render(self, sample_pdf):
        from src.pdf_preview import get_page_count, render_page
        count = get_page_count(sample_pdf)
        _, total = render_page(sample_pdf, page=0)
        assert count == total

    def test_single_page_pdf(self, sample_pdf):
        from src.pdf_preview import get_page_count
        count = get_page_count(sample_pdf)
        assert count >= 1
